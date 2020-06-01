#!/usr/bin/env python3
"""
Output
"""

import json
import csv
import os
import sqlite3 as db
from openpyxl import Workbook

#mencari home direktori
def home():
    """cari alamat HOME"""
    if os.name == 'nt':
        try:
            result = os.path.join(os.environ["HOMEPATH"], "My Documents")
        except Exception:
            result = os.environ["HOMEDRIVE"]
    else:
        result = os.path.join(os.environ["HOME"], "user")
    return result

#fungsi ouput dengan 5 pilihan : layar, excell, JSON, csv, text

def pt_screen(buffer):
    """output ke layar"""
    print(buffer)

def pt_excell(file_title, buffer):
    """output ke excell"""

    #membaca data header dari judul.json
    with open('data/judul.json', 'r') as head_data:
        header = list(json.load(head_data).keys())

    book = Workbook()
    sheet = book.active
    sheet.title = file_title

    #header untuk mengisi baris pertama
    for i in range(len(header)):
        sheet.cell(column=i + 1, row=1, value=header[i])

    #isi sheet dari buffer
    for rows in range(len(buffer)):
        for cols in range(len(buffer[rows])):
            sheet.cell(column=cols + 1, row=rows+2, value=buffer[rows][str(cols)])

    #save ke excell
    book.save('{}.xlsx'.format(home() + file_title))
    book.close()
    print(f'{file_title} berhasil dicetak ke {home()}/{file_title}.xlsx ')

def pt_json(file_title, buffer):
    """output ke json"""
    with open('{}.json'.format(home() + file_title), 'w') as jsonfile:
        data = json.dumps(buffer)
        jsonfile.write(data)
    print(f'{file_title} berhasil dicetak ke {home()}/{file_title}.json ')

def pt_csv(file_title, buffer):
    """output ke csv"""
    with open(f'{home()+file_title}.csv', 'w', newline='') as csvfile:
        x = csv.writer(csvfile, delimiter=',', quotechar='"')
        for i in buffer:
            x.writerow(i)
    print(f'{file_title} berhasil dicetak ke {home()}/{file_title}.csv ')

def pt_txt(file_title, buffer):
    """output ke text"""
    with open('{}.txt'.format(home() + file_title), 'w') as txtfile:
        txtfile.write(json.dumps(buffer))
    print(f'{file_title} berhasil dicetak ke {home()}/{file_title}.txt ')

def pt_db(buffer):
    """simpan ke database"""
    conn = db.connect("data/report.db")
    for rows in buffer:
        conn.execute(f'''insert into laporan(
`NoPeg`
,`No. Akun`
,`No.`
,`Nama`
,`Auto-Assign`
,`Tanggal`
,`Jam Kerja`
,`Awal tugas`
,`Akhir tugas`
,`Masuk`
,`Keluar`
,`Normal`
,`Waktu real`
,`Telat`
,`Plg Awal`
,`Bolos`
,`Waktu Lembur`
,`Waktu Kerja`
,`Status`
,`Hrs C/In`
,`Hrs C/Out`
,`Departemen`
,`NDays`
,`Akhir Pekan`
,`Hari Libur`
,`Lama Hadir`
,`NDays_OT`
,`Lembur A.Pekan`
,`Libur Lembur`)
 values("{rows['0']}"
,"{rows['1']}"
,"{rows['2']}"
,"{rows['3']}"
,"{rows['4']}"
,"{rows['5']}"
,"{rows['6']}"
,"{rows['7']}"
,"{rows['8']}"
,"{rows['9']}"
,"{rows['10']}"
,"{rows['11']}"
,"{rows['12']}"
,"{rows['13']}"
,"{rows['14']}"
,"{rows['15']}"
,"{rows['16']}"
,"{rows['17']}"
,"{rows['18']}"
,"{rows['19']}"
,"{rows['20']}"
,"{rows['21']}"
,"{rows['22']}"
,"{rows['23']}"
,"{rows['24']}"
,"{rows['25']}"
,"{rows['26']}"
,"{rows['27']}"
,"{rows['28']}");''')
    conn.commit()
    conn.close()

def main():
    "modul utama"
    print("not ready yet")

if __name__ == '__main__':
    main()
