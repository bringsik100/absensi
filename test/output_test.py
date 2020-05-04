#!/usr/bin/env python3
"""
Output
"""


import json
import csv
import string
from openpyxl import Workbook
import os
import pytest

"""mencari home direktori"""
def test_home():
	if os.name == 'nt':
		return os.environ["HOME"]+"/My Documents/"
	else:
		return os.environ["HOME"]+"/user/"

"""fungsi ouput dengan 5 pilihan : layar, excell, JSON, csv, text"""

def test_pt_screen(buffer):
	"""output ke layar"""
	print(buffer)
		

def test_pt_excell(file_title,buffer):
	"""output ke excell"""

	"""membaca data header dari judul.json"""
	with open('data/judul.json','r') as head_data:
		header = list(json.load(head_data).keys())

	book = Workbook()
	sheet = book.active
	sheet.title = file_title

	"""header untuk mengisi baris pertama"""
	for i in range(len(header)):
		sheet.cell(column=i+1,row=1, value=header[i])

	"""isi sheet dari buffer"""
	for rows in range(len(buffer)):
		for cols in range(len((buffer[rows]))):
			sheet.cell(column=cols+1,row=rows+2,value=buffer[rows][str(cols)])
		
	"""save ke excell"""
	book.save('{}.xlsx'.format(os.path.join(test_home(),file_title)))
	book.close()

def test_pt_json(file_title,buffer):
	"""output ke json"""
	with open('{}.json'.format(os.path.join(test_home(),file_title)),'w') as jsonfile:
		json.dumps(buffer)

def test_pt_csv(file_title,buffer):
	"""output ke csv"""
	with open(f'{test_home()+file_title}.csv', 'w', newline='') as csvfile:
		x = csv.writer(csvfile, delimiter=',',quotechar='"')
		for i in buffer:
			x.writerow(i)

def test_pt_txt(file_title,buffer):
	"""output ke text"""
	with open('{}.txt'.format(test_home()+file_title), 'w') as txtfile:
			txtfile.write(json.dumps(buffer))

def test_main():
	print("not ready yet")

if __name__ == '__main__':
	test_main()
